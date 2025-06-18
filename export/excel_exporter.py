import pandas as pd
from datetime import datetime
import tkinter as tk
from tkinter import filedialog
from core.app_state import APP_STATE
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils import get_column_letter


def export_to_excel():
    try:
        if not APP_STATE.get("dados_extraidos"):
            return {"success": False, "message": "Nenhum dado disponível para exportação"}

        ordem_colunas = APP_STATE.get("ordem_colunas", [])

        todos_dados = []
        for doc_id, dados_list in APP_STATE["dados_extraidos"].items():
            if dados_list:
                todos_dados.extend(dados_list)

        if not todos_dados:
            return {"success": False, "message": "Nenhum dado válido encontrado"}

        df = pd.DataFrame(todos_dados)
        colunas_finais = [col for col in ordem_colunas if col in df.columns]
        df = df[colunas_finais]

        root = tk.Tk()
        root.withdraw()

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_filename = f"consorcios_exportados_{timestamp}.xlsx"

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")],
            initialfile=default_filename,
            title="Salvar arquivo Excel como..."
        )

        if not filepath:
            return {"success": False, "message": "Exportação cancelada pelo usuário"}

        # Exporta para Excel (sem index)
        df.to_excel(filepath, index=False, engine='openpyxl')

        # Reabre com openpyxl para formatar
        wb = load_workbook(filepath)
        ws = wb.active

        # Formatação do cabeçalho
        header_fill = PatternFill(start_color="CC092F", end_color="CC092F", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_num, col in enumerate(ws.iter_cols(min_row=1, max_row=1), 1):
            for cell in col:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

        # Auto filtro na primeira linha
        max_col = ws.max_column
        max_row = ws.max_row
        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

        # Formata colunas monetárias
        colunas_monetarias = [
            "Valor Crédito", "Líquido a Pagar", "Saldo Devedor",
            "Val. Disponíveis Encerramento", "Valor Contrib. Mensal"
        ]
        formato_monetario = u'R$ #,##0.00'

        for col_idx, col_name in enumerate(df.columns, 1):
            if col_name in colunas_monetarias:
                for row in range(2, max_row + 1):  # começa na linha 2 (dados)
                    cell = ws.cell(row=row, column=col_idx)
                    try:
                        valor = float(str(cell.value).replace("R$", "").replace(".", "").replace(",", "."))
                        cell.value = valor
                        cell.number_format = formato_monetario
                    except:
                        pass  # ignora se não for número

        for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=max_row), 1):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            for cell in col:
                try:
                    cell_value = str(cell.value) if cell.value is not None else ""
                    max_length = max(max_length, len(cell_value))
                except:
                    pass
            # Ajuste de largura com margem extra
            adjusted_width = max_length + 2
            ws.column_dimensions[col_letter].width = adjusted_width


        wb.save(filepath)

        return {
            "success": True,
            "message": f"Arquivo salvo com sucesso em:\n{filepath}",
            "filepath": filepath
        }

    except Exception as e:
        return {
            "success": False,
            "message": f"Erro durante exportação: {str(e)}"
        }
